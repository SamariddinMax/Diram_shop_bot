from __future__ import annotations
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.filters import Command, CommandStart
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
from aiogram import Bot, Dispatcher, F
from passbot import ADMIN_ID, TOKEN
from typing import Optional
from aiogram.types import CallbackQuery, InlineKeyboardButton, InlineKeyboardMarkup, KeyboardButton, Message, ReplyKeyboardMarkup
import aiosqlite
import asyncio
from pathlib import Path
import csv
import io
import sqlite3

DB_PATH = "shop.db"
USERS_CSV = Path("users.csv")

bot = Bot(TOKEN)
dp  = Dispatcher()

# ===================== FSM STATES =====================

class AdminStates(StatesGroup):
    waiting_price_key                = State()
    waiting_price_value              = State()
    waiting_text_key                 = State()
    waiting_text_value               = State()
    waiting_new_product_key          = State()
    waiting_new_product_category     = State()
    waiting_new_product_name         = State()
    waiting_new_product_price        = State()
    waiting_new_product_photo        = State()
    waiting_file_import              = State()
    waiting_photo_key                = State()
    waiting_photo_new                = State()
    waiting_new_category_name        = State()
    waiting_category_rename_select   = State()
    waiting_category_rename_new      = State()
    waiting_category_delete_select   = State()
    waiting_category_delete_confirm  = State()
    waiting_delete_product           = State()
    # FIX #6: Added missing broadcast FSM state
    waiting_broadcast_text           = State()


class DebtState(StatesGroup):
    phone   = State()
    confirm = State()
    amount  = State()


class DeleteDebt(StatesGroup):
    phone  = State()
    amount = State()


class RegisterState(StatesGroup):
    phone = State()

# ===================== DATABASE SETUP =====================

async def init_db() -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.executescript("""
            CREATE TABLE IF NOT EXISTS products (
                key      TEXT PRIMARY KEY,
                name     TEXT NOT NULL,
                price    REAL NOT NULL DEFAULT 0,
                photo_id TEXT DEFAULT NULL,
                category TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS texts (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL DEFAULT ''
            );

            CREATE TABLE IF NOT EXISTS cart (
                user_id     INTEGER NOT NULL,
                product_key TEXT    NOT NULL,
                qty         INTEGER NOT NULL DEFAULT 1,
                PRIMARY KEY (user_id, product_key),
                FOREIGN KEY (product_key) REFERENCES products(key)
            );

            CREATE TABLE IF NOT EXISTS debts (
                phone      TEXT PRIMARY KEY,
                amount     REAL NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL DEFAULT (datetime('now'))
            );

            CREATE TABLE IF NOT EXISTS users (
                phone       TEXT PRIMARY KEY,
                telegram_id INTEGER NOT NULL
            );

            CREATE TABLE IF NOT EXISTS categories (
               id   INTEGER PRIMARY KEY AUTOINCREMENT,
               name TEXT UNIQUE NOT NULL
            );

            INSERT OR IGNORE INTO categories(name) VALUES ('drinks');
            INSERT OR IGNORE INTO categories(name) VALUES ('sweets');
            INSERT OR IGNORE INTO categories(name) VALUES ('ice');

            CREATE TABLE IF NOT EXISTS orders (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id      INTEGER NOT NULL,
                user_name    TEXT    NOT NULL DEFAULT '',
                phone        TEXT    NOT NULL DEFAULT '',
                items_text   TEXT    NOT NULL DEFAULT '',
                total        REAL    NOT NULL DEFAULT 0,
                status       TEXT    NOT NULL DEFAULT 'pending',
                created_at   TEXT    NOT NULL DEFAULT (datetime('now'))
            );
        """)

        # Migrate products table
        for col, definition in [
            ("photo_id", "TEXT DEFAULT NULL"),
            ("category", "TEXT NOT NULL DEFAULT ''"),
        ]:
            try:
                await db.execute(f"ALTER TABLE products ADD COLUMN {col} {definition}")
                await db.commit()
            except Exception:
                pass

        # Migrate orders table
        for col, definition in [
            ("user_name",  "TEXT NOT NULL DEFAULT ''"),
            ("phone",      "TEXT NOT NULL DEFAULT ''"),
            ("items_text", "TEXT NOT NULL DEFAULT ''"),
            ("total",      "REAL NOT NULL DEFAULT 0"),
            ("status",     "TEXT NOT NULL DEFAULT 'pending'"),
            ("created_at", "TEXT NOT NULL DEFAULT (datetime('now'))"),
        ]:
            try:
                await db.execute(f"ALTER TABLE orders ADD COLUMN {col} {definition}")
                await db.commit()
            except Exception:
                pass

        seed_products: list[tuple[str, str, float, str]] = []
        for key, name, price, cat in seed_products:
            await db.execute(
                "INSERT OR IGNORE INTO products (key, name, price, category) VALUES (?, ?, ?, ?)",
                (key, name, price, cat),
            )

        seed_texts: list[tuple[str, str]] = [
            ("start",      "Хуш омадед! Категорияро интихоб кунед."),
            ("cart_empty", "🛒 Сабади шумо холӣ аст."),
            ("contact",    "📞 Барои тамос: @admin"),
            ("discount",   "🎁 Тахфифи имрӯза нест."),
        ]
        for key, value in seed_texts:
            await db.execute(
                "INSERT OR IGNORE INTO texts (key, value) VALUES (?, ?)",
                (key, value),
            )

        await db.commit()

# ===================== CATEGORY HELPERS (sync) =====================

def get_categories() -> list[str]:
    """Return list of category names ordered by id."""
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS categories (
            id   INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE NOT NULL
        )
        """
    )
    conn.commit()
    for d in ("drinks", "sweets", "ice"):
        try:
            cur.execute("INSERT OR IGNORE INTO categories(name) VALUES (?)", (d,))
        except Exception:
            pass
    conn.commit()
    cur.execute("SELECT name FROM categories ORDER BY id")
    rows = cur.fetchall()
    conn.close()
    return [row[0] for row in rows]


def add_category_sync(name: str) -> bool:
    """Add a new category. Returns True if inserted, False if duplicate/invalid."""
    name = (name or "").strip()
    if not name:
        return False
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    try:
        cur.execute("INSERT INTO categories(name) VALUES (?)", (name,))
        conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False
    finally:
        conn.close()


def build_category_kb(one_time: bool = True) -> ReplyKeyboardMarkup:
    """Build a ReplyKeyboardMarkup dynamically from categories in DB."""
    cats: list[str] = get_categories()
    rows: list[list[KeyboardButton]] = []
    for i in range(0, len(cats), 2):
        row = [KeyboardButton(text=str(c)) for c in cats[i: i + 2]]
        rows.append(row)
    rows.append([KeyboardButton(text="⬅️ Бозгашт")])
    return ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True, one_time_keyboard=one_time)


def add_user(user_id: int) -> None:
    """Илова кардани корбар ба users.csv агар набошад."""
    users = get_all_users()
    if user_id not in users:
        with open(USERS_CSV, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([user_id])


def get_all_users() -> list[int]:
    """Хондани ҳама корбарон аз файл."""
    # FIX #1: was using sqlite3.Row instead of iterating CSV rows
    if not USERS_CSV.exists():
        return []
    result: list[int] = []
    with open(USERS_CSV, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row in reader:
            if row:
                try:
                    result.append(int(row[0]))
                except (ValueError, IndexError):
                    pass
    return result


# ===================== CATEGORY DB HELPERS (async) =====================

async def db_get_categories() -> list[str]:
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT name FROM categories ORDER BY id") as cur:
            rows = await cur.fetchall()
    return [row[0] for row in rows]


async def db_add_category(name: str) -> bool:
    name = name.strip().lower()
    if not name:
        return False
    try:
        async with aiosqlite.connect(DB_PATH) as db:
            await db.execute("INSERT INTO categories(name) VALUES(?)", (name,))
            await db.commit()
        return True
    except aiosqlite.IntegrityError:
        return False


async def db_rename_category(old_name: str, new_name: str) -> bool:
    """Rename category and update all products. Returns False if new name already exists."""
    async with aiosqlite.connect(DB_PATH) as db:
        try:
            await db.execute("BEGIN")
            cur = await db.execute(
                "SELECT 1 FROM categories WHERE LOWER(name)=LOWER(?)", (new_name,)
            )
            exists = await cur.fetchone()
            if exists:
                await db.rollback()
                return False
            await db.execute(
                "UPDATE categories SET name=? WHERE name=?", (new_name, old_name)
            )
            await db.execute(
                "UPDATE products SET category=? WHERE category=?", (new_name, old_name)
            )
            await db.commit()
            return True
        except Exception:
            await db.rollback()
            raise


async def db_delete_category(name: str) -> bool:
    async with aiosqlite.connect(DB_PATH) as db:
        # Ensure "other" category exists before reassigning products to it
        await db.execute("INSERT OR IGNORE INTO categories(name) VALUES ('other')")
        # Move products to "other" instead of deleting them
        await db.execute(
            "UPDATE products SET category='other' WHERE category=?",
            (name,)
        )
        await db.execute(
            "DELETE FROM categories WHERE name=?",
            (name,)
        )
        await db.commit()
        return True


# ===================== DB HELPERS =====================

async def db_get_product(key: str) -> Optional[dict[str, object]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM products WHERE key = ?",
            (key,)
        ) as cur:
            row = await cur.fetchone()
            return dict(row) if row else None


async def db_all_products() -> list[dict[str, object]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM products ORDER BY name") as cur:
            return [dict(r) for r in await cur.fetchall()]


async def db_products_by_category(category: str) -> list[dict[str, object]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT * FROM products WHERE category = ? ORDER BY name", (category,)
        ) as cur:
            return [dict(r) for r in await cur.fetchall()]


async def db_upsert_product(
    key: str,
    name: str,
    price: float,
    photo_id: Optional[str] = None,
    category: str = "other",
) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO products (key, name, price, photo_id, category) VALUES (?, ?, ?, ?, ?) "
            "ON CONFLICT(key) DO UPDATE SET name=excluded.name, price=excluded.price, "
            "photo_id=COALESCE(excluded.photo_id, products.photo_id), "
            "category=excluded.category",
            (key, name, price, photo_id, category),
        )
        await db.commit()


async def db_set_product_photo(key: str, photo_id: str) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE products SET photo_id = ? WHERE key = ?", (photo_id, key)
        )
        await db.commit()


async def db_get_text(key: str) -> str:
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute("SELECT value FROM texts WHERE key = ?", (key,)) as cur:
            row = await cur.fetchone()
            return str(row[0]) if row else ""


async def db_all_texts() -> list[dict[str, object]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM texts") as cur:
            return [dict(r) for r in await cur.fetchall()]


async def db_set_text(key: str, value: str) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO texts (key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )
        await db.commit()


async def db_cart_add(user_id: int, product_key: str) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO cart (user_id, product_key, qty) VALUES (?, ?, 1) "
            "ON CONFLICT(user_id, product_key) DO UPDATE SET qty = qty + 1",
            (user_id, product_key),
        )
        await db.commit()


async def db_cart_get(user_id: int) -> list[dict[str, object]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT c.product_key, c.qty, p.name, p.price "
            "FROM cart c JOIN products p ON c.product_key = p.key "
            "WHERE c.user_id = ?",
            (user_id,),
        ) as cur:
            return [dict(r) for r in await cur.fetchall()]


async def db_cart_clear(user_id: int) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM cart WHERE user_id = ?", (user_id,))
        await db.commit()


# ===================== ORDER DB HELPERS =====================

async def db_create_order(user_id: int, user_name: str, phone: str, items_text: str, total: float) -> int:
    """Save a new order and return its auto-increment id."""
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "INSERT INTO orders (user_id, user_name, phone, items_text, total) VALUES (?, ?, ?, ?, ?)",
            (user_id, user_name, phone, items_text, total),
        )
        await db.commit()
        return cur.lastrowid  # type: ignore[return-value]


async def db_get_order(order_id: int) -> Optional[dict[str, object]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute("SELECT * FROM orders WHERE id = ?", (order_id,)) as cur:
            row = await cur.fetchone()
            return dict(row) if row else None


async def db_update_order_status(order_id: int, status: str) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("UPDATE orders SET status=? WHERE id=?", (status, order_id))
        await db.commit()


async def db_delete_product(key: str) -> bool:
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "DELETE FROM products WHERE key = ?",
            (key,)
        )
        await db.commit()
        return cur.rowcount > 0

# ===================== DEBT DB HELPERS =====================

async def db_get_debt(phone: str) -> Optional[dict[str, object]]:
    async with aiosqlite.connect(DB_PATH) as db:
        db.row_factory = aiosqlite.Row
        async with db.execute(
            "SELECT phone, amount, created_at FROM debts WHERE phone = ?", (phone,)
        ) as cur:
            row = await cur.fetchone()
            return dict(row) if row else None


async def db_create_debt(phone: str, amount: float) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO debts (phone, amount) VALUES (?, ?)", (phone, amount)
        )
        await db.commit()


async def db_update_debt(phone: str, new_amount: float) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE debts SET amount = ? WHERE phone = ?", (new_amount, phone)
        )
        await db.commit()


async def db_delete_debt(phone: str) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("DELETE FROM debts WHERE phone = ?", (phone,))
        await db.commit()


async def db_get_user_telegram_id(phone: str) -> Optional[int]:
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            "SELECT telegram_id FROM users WHERE phone = ?", (phone,)
        ) as cur:
            row = await cur.fetchone()
            return int(row[0]) if row else None


async def db_get_phone_by_telegram_id(telegram_id: int) -> Optional[str]:
    """Reverse lookup: get registered phone for a given Telegram user ID."""
    async with aiosqlite.connect(DB_PATH) as db:
        async with db.execute(
            "SELECT phone FROM users WHERE telegram_id = ?", (telegram_id,)
        ) as cur:
            row = await cur.fetchone()
            return str(row[0]) if row else None


async def db_register_user(phone: str, telegram_id: int) -> None:
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO users (phone, telegram_id) VALUES (?, ?) "
            "ON CONFLICT(phone) DO UPDATE SET telegram_id=excluded.telegram_id",
            (phone, telegram_id),
        )
        await db.commit()


async def db_cart_get_by_phone(phone: str) -> list[dict[str, object]]:
    telegram_id = await db_get_user_telegram_id(phone)
    if telegram_id is None:
        return []
    return await db_cart_get(telegram_id)


async def db_save_telegram_id(telegram_id: int) -> None:
    """Persist a telegram_id to the bot_users table (no phone required)."""
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "CREATE TABLE IF NOT EXISTS bot_users "
            "(telegram_id INTEGER PRIMARY KEY)"
        )
        await db.execute(
            "INSERT OR IGNORE INTO bot_users(telegram_id) VALUES(?)",
            (telegram_id,),
        )
        await db.commit()


async def db_get_all_telegram_ids() -> list[int]:
    """Return every known telegram_id from bot_users + registered users, deduplicated."""
    ids: set[int] = set()
    async with aiosqlite.connect(DB_PATH) as db:
        # Registered users (phone → telegram_id)
        async with db.execute("SELECT telegram_id FROM users") as cur:
            for row in await cur.fetchall():
                ids.add(int(row[0]))
        # All users who ever pressed /start (no phone required)
        try:
            async with db.execute("SELECT telegram_id FROM bot_users") as cur:
                for row in await cur.fetchall():
                    ids.add(int(row[0]))
        except Exception:
            pass
    # Also include anyone saved in users.csv as a last-resort fallback
    for uid in get_all_users():
        ids.add(uid)
    return list(ids)


def is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID

# ===================== FILE IMPORT (CSV / Excel) =====================

def parse_import_file(data: bytes, filename: str) -> list[dict[str, object]] | str:
    try:
        rows: list[dict[str, object]] = []

        if filename.endswith(".csv"):
            text   = data.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows   = [dict(r) for r in reader]

        elif filename.endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data))
            ws = wb.active
            if ws is None:
                return "❌ Файл холӣ аст ё варақи фаъол нест."
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            headers    = [str(cell.value).strip().lower() for cell in header_row]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(
                    {headers[i]: ("" if v is None else v) for i, v in enumerate(row)}
                )
        else:
            return "❌ Формати файл дастгирӣ намешавад. Танҳо .csv ё .xlsx фиристед."

        valid_cats = set(get_categories())
        products: list[dict[str, object]] = []
        for i, row in enumerate(rows, start=2):
            key       = str(row.get("key", "")).strip().lower().replace(" ", "_")[:20]
            name      = str(row.get("name", "")).strip()
            price_raw = str(row.get("price", "0")).strip()
            if not key or not name:
                continue
            try:
                price = float(price_raw)
            except ValueError:
                return f"❌ Хато дар сатри {i}: нарх рақам нест — «{price_raw}»"
            photo_val    = str(row.get("photo_id", "")).strip()
            category_val = str(row.get("category", "")).strip().lower()
            if not category_val or category_val == "none" or category_val not in valid_cats:
                category_val = "other"
            products.append({
                "key":      key,
                "name":     name,
                "price":    price,
                "photo_id": photo_val if photo_val else None,
                "category": category_val,
            })

        if not products:
            return "❌ Файл холӣ аст ё сутунҳо дуруст нестанд (key, name, price лозим аст)."
        return products

    except Exception as exc:
        return f"❌ Хато ҳангоми хондани файл: {exc}"

# ===================== SHARED HELPERS =====================

def _cart_text(items: list[dict[str, object]]) -> tuple[str, float]:
    lines: list[str] = []
    total = 0.0
    for item in items:
        subtotal = float(item["price"]) * int(item["qty"])  # type: ignore
        total   += subtotal
        lines.append(f"  • {item['name']} × {item['qty']} = {subtotal:.2f} сомонӣ")
    return "\n".join(lines), total


async def _notify_user(telegram_id: int, text: str, parse_mode: str = "HTML") -> None:
    try:
        await bot.send_message(telegram_id, text, parse_mode=parse_mode)
    except Exception:
        pass

# ===================== KEYBOARDS =====================

main_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="🛍 Маҳсулот"), KeyboardButton(text="🛒 Сабад")],
        [KeyboardButton(text="📞 Алоқа"),    KeyboardButton(text="🎁 Тахфиф")],
    ],
    resize_keyboard=True,
)

admin_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="💰 Иваз кардани нарх"),  KeyboardButton(text="📝 Иваз кардани матн")],
        [KeyboardButton(text="➕ Маҳсулоти нав"),      KeyboardButton(text="🗑 Нархҳоро нишон диҳед")],
        [KeyboardButton(text="📥 Аз файл илова кун"),  KeyboardButton(text="💼 Иваз кардани расм")],
        [KeyboardButton(text="➕ Иловаи қарз"),        KeyboardButton(text="❌ Кам кардани қарз")],
        [KeyboardButton(text="📂 Идораи категорияхо"), KeyboardButton(text="🗑 Нест кардани маҳсулот")],
        [KeyboardButton(text="📨 Хабардиҳӣ"),          KeyboardButton(text="⬅️ Бозгашт")],
    ],
    resize_keyboard=True,
)

category_manage_kb = ReplyKeyboardMarkup(
    keyboard=[
        [KeyboardButton(text="➕ Илова кардани категория")],
        [KeyboardButton(text="✏️ Тағйири номи категория")],
        [KeyboardButton(text="🗑 Нест кардани категория")],
        [KeyboardButton(text="📋 Рӯйхати категорияҳо")],
        [KeyboardButton(text="⬅️ Бозгашт")],
    ],
    resize_keyboard=True,
)

# ===================== START / BACK =====================

@dp.message(CommandStart())
async def start(message: Message) -> None:
    user = message.from_user
    if user:
        add_user(user.id)                    # keep CSV for backwards-compat
        await db_save_telegram_id(user.id)   # also persist in DB for broadcast
    if user and is_admin(user.id):
        await message.answer("👨‍💼 Панели админ:", reply_markup=admin_kb)
    else:
        await message.answer("🏠 Менюи асосӣ", reply_markup=main_kb)


@dp.message(F.text == "⬅️ Бозгашт")
async def back(message: Message, state: FSMContext) -> None:
    await state.clear()
    user = message.from_user
    if user and is_admin(user.id):
        await message.answer("👨‍💼 Панели админ:", reply_markup=admin_kb)
    else:
        await message.answer("🏠 Менюи асосӣ", reply_markup=main_kb)


# ===================== USER PHONE REGISTRATION =====================

@dp.message(Command("register"))
async def register_start(message: Message, state: FSMContext) -> None:
    await message.answer(
        "📱 Рақами телефони худро нависед.\n"
        "Намуна: <code>+992901234567</code>",
        parse_mode="HTML",
    )
    await state.set_state(RegisterState.phone)


@dp.message(RegisterState.phone, F.text)
async def register_phone(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None:
        return
    phone = message.text.strip()  # type: ignore[union-attr]
    await db_register_user(phone, user.id)
    await state.clear()
    await message.answer(
        f"✅ Рақами <code>{phone}</code> ба ҳисоби шумо пайваст шуд.",
        parse_mode="HTML",
    )

# ===================== PRODUCTS =====================

CATEGORY_EMOJI_MAP: dict[str, str] = {
    "drinks": "🥤 Нӯшокиҳо",
    "sweets": "🍰 Шириниҳо",
    "ice":    "🍨 Яхмосҳо",
}

def _cat_label(key: str) -> str:
    return CATEGORY_EMOJI_MAP.get(key, key.capitalize())

# FIX #3: Dynamic categories now resolve correctly from plain text labels
def _cat_key_from_label(label: str) -> Optional[str]:
    # First check the hardcoded emoji map (reverse lookup)
    for key, lbl in CATEGORY_EMOJI_MAP.items():
        if lbl == label:
            return key
    # Then check plain category names (for dynamically added categories)
    cats = get_categories()
    label_lower = label.strip().lower()
    for cat in cats:
        if cat.lower() == label_lower or cat.capitalize() == label:
            return cat
    return None


@dp.message(F.text == "🛍 Маҳсулот")
async def products(message: Message) -> None:
    cats = get_categories()
    visible = [c for c in cats if c != "other"]
    rows: list[list[KeyboardButton]] = []
    for i in range(0, len(visible), 2):
        row = [
            KeyboardButton(text=_cat_label(visible[j]))
            for j in range(i, min(i + 2, len(visible)))
        ]
        rows.append(row)
    rows.append([KeyboardButton(text="⬅️ Бозгашт")])
    kb = ReplyKeyboardMarkup(keyboard=rows, resize_keyboard=True)
    await message.answer("📦 Категорияро интихоб кунед:", reply_markup=kb)


async def _send_category_menu(message: Message, category: str, title: str) -> None:
    prods = await db_products_by_category(category)
    if not prods:
        await message.answer(f"{title}\n\n⚠️ Ҳоло маҳсулот вуҷуд надорад.")
        return
    rows = []
    for i in range(0, len(prods), 2):
        row = [
            InlineKeyboardButton(text=str(p["name"]), callback_data=str(p["key"]))
            for p in prods[i: i + 2]
        ]
        rows.append(row)
    kb = InlineKeyboardMarkup(inline_keyboard=rows)
    await message.answer(f"{title}:", reply_markup=kb)


# FIX #4: Kept specific handlers for the 3 built-in categories AND
# the catch-all dynamic_category_handler below covers admin-added ones.
@dp.message(F.text == "🥤 Нӯшокиҳо")
async def drinks(message: Message) -> None:
    await _send_category_menu(message, "drinks", "🥤 Нӯшокиро интихоб кунед")


@dp.message(F.text == "🍰 Шириниҳо")
async def sweets(message: Message) -> None:
    await _send_category_menu(message, "sweets", "🍰 Шириниро интихоб кунед")


@dp.message(F.text == "🍨 Яхмосҳо")
async def ice(message: Message) -> None:
    await _send_category_menu(message, "ice", "🍨 Яхмосро интихоб кунед")


# ===================== CALLBACKS =====================

@dp.callback_query()
async def handle_callback(callback: CallbackQuery) -> None:
    data = callback.data
    if not data:
        await callback.answer()
        return

    if callback.message is None:
        await callback.answer("❌ Хатогӣ рух дод.", show_alert=True)
        return

    if data == "clear_cart":
        await db_cart_clear(callback.from_user.id)
        await callback.message.answer("🛒 Сабади шумо тоза шуд.")
        await callback.answer()
        return

    if data == "send_order":
        items = await db_cart_get(callback.from_user.id)

        if not items:
            await callback.answer("🛒 Сабад холӣ аст.", show_alert=True)
            return

        total = 0.0
        user  = callback.from_user
        lines: list[str] = []

        for item in items:
            subtotal = float(item["price"]) * int(item["qty"])  # type: ignore
            total   += subtotal
            lines.append(
                f"  • {item['name']} × {item['qty']} = {subtotal:.2f} сомонӣ"
            )

        items_text = "\n".join(lines)
        phone = await db_get_phone_by_telegram_id(user.id) or "—"

        if user.username:
            username_line = f"🔗 Username: @{user.username}"
        else:
            username_line = "🔗 Username: <i>нест</i>"

        order_id = await db_create_order(
            user_id    = user.id,
            user_name  = user.full_name or str(user.id),
            phone      = phone,
            items_text = items_text,
            total      = total,
        )

        card = (
            f"🛍 <b>ФАРМОИШИ НАВ</b>  #<b>{order_id}</b>\n\n"
            f"👤 Ном: <b>{user.full_name}</b>\n"
            f"🆔 ID: <code>{user.id}</code>\n"
            f"{username_line}\n"
            f"📱 Телефон: <code>{phone}</code>\n\n"
            f"📦 Маҳсулот:\n{items_text}\n\n"
            f"💰 <b>Ҷамъ: {total:.2f} сомонӣ</b>"
        )

        admin_order_kb = InlineKeyboardMarkup(
            inline_keyboard=[
                [
                    InlineKeyboardButton(
                        text="✅ Тасдиқи фармоиш",
                        callback_data=f"order_confirm:{order_id}",
                    ),
                    InlineKeyboardButton(
                        text="❌ Рад кардан",
                        callback_data=f"order_reject:{order_id}",
                    ),
                ],
                [
                    InlineKeyboardButton(
                        text="💬 Навиштан ба харидор",
                        url=f"tg://user?id={user.id}",
                    ),
                ],
            ]
        )

        await bot.send_message(ADMIN_ID, card, parse_mode="HTML", reply_markup=admin_order_kb)
        await db_cart_clear(user.id)
        await callback.message.answer(
            f"✅ Фармоиши шумо (#<b>{order_id}</b>) ба админ фиристода шуд.\n"
            f"⏳ Лутфан интизор шавед, АДМИН тасдиқ мекунад.",
            parse_mode="HTML",
        )
        await callback.answer()
        return

    if data.startswith("order_confirm:"):
        if not is_admin(callback.from_user.id):
            await callback.answer("⛔ Танҳо барои админ.", show_alert=True)
            return
        order_id = int(data.split(":")[1])
        order    = await db_get_order(order_id)
        if order is None:
            await callback.answer("❌ Фармоиш ёфт нашуд.", show_alert=True)
            return
        if str(order["status"]) != "pending":
            await callback.answer("⚠️ Ин фармоиш аллакай коркард шудааст.", show_alert=True)
            return

        await db_update_order_status(order_id, "confirmed")

        confirmed_text = (
            f"✅ <b>ТАСДИҚ ШУД</b>  #<b>{order_id}</b>\n\n"
            f"👤 {order['user_name']}  (<code>{order['user_id']}</code>)\n"
            f"📱 Телефон: <code>{order['phone']}</code>\n\n"
            f"📦 Маҳсулот:\n{order['items_text']}\n\n"
            f"💰 <b>Ҷамъ: {float(order['total']):.2f} сомонӣ</b>\n\n"  # type: ignore
            f"🕐 {order['created_at']}"
        )
        await callback.message.edit_text(confirmed_text, parse_mode="HTML")  # type: ignore[union-attr]

        await _notify_user(
            int(order["user_id"]),  # type: ignore
            f"🎉 Фармоиши шумо (#<b>{order_id}</b>) тасдиқ шуд!\n\n"
            f"📦 Маҳсулот:\n{order['items_text']}\n\n"
            f"💰 <b>Ҷамъ: {float(order['total']):.2f} сомонӣ</b>\n\n"  # type: ignore
            "📞 Барои тафсилот бо мо тамос гиред.",
        )
        await callback.answer("✅ Фармоиш тасдиқ шуд!")
        return

    if data.startswith("order_reject:"):
        if not is_admin(callback.from_user.id):
            await callback.answer("⛔ Танҳо барои админ.", show_alert=True)
            return
        order_id = int(data.split(":")[1])
        order    = await db_get_order(order_id)
        if order is None:
            await callback.answer("❌ Фармоиш ёфт нашуд.", show_alert=True)
            return
        if str(order["status"]) != "pending":
            await callback.answer("⚠️ Ин фармоиш аллакай коркард шудааст.", show_alert=True)
            return

        await db_update_order_status(order_id, "rejected")

        rejected_text = (
            f"❌ <b>РАД КАРДА ШУД</b>  #<b>{order_id}</b>\n\n"
            f"👤 {order['user_name']}  (<code>{order['user_id']}</code>)\n"
            f"📱 Телефон: <code>{order['phone']}</code>\n\n"
            f"📦 Маҳсулот:\n{order['items_text']}\n\n"
            f"💰 <b>Ҷамъ: {float(order['total']):.2f} сомонӣ</b>\n\n"  # type: ignore
            f"🕐 {order['created_at']}"
        )
        await callback.message.edit_text(rejected_text, parse_mode="HTML")  # type: ignore[union-attr]

        await _notify_user(
            int(order["user_id"]),  # type: ignore
            f"😔 Мутаассифона, фармоиши шумо (#<b>{order_id}</b>) рад карда шуд.\n\n"
            "📞 Барои тафсилот бо мо тамос гиред.",
        )
        await callback.answer("❌ Фармоиш рад карда шуд.")
        return

    # Product key callback — add to cart
    product = await db_get_product(data)
    if product is None:
        await callback.answer("❌ Маҳсулот ёфт нашуд.", show_alert=True)
        return

    await db_cart_add(callback.from_user.id, data)
    await callback.answer(
        f"✅ {product['name']} ба сабад илова шуд!",
        show_alert=False,
    )


@dp.message(F.text == "🗑 Нест кардани маҳсулот")
async def delete_product_start(message: Message, state: FSMContext):
    prods = await db_all_products()

    if not prods:
        await message.answer("❌ Дар база ягон маҳсулот вуҷуд надорад.")
        return

    text = "\n".join(
        f"• <code>{p['key']}</code> — {p['name']}"
        for p in prods
    )

    await message.answer(
        "Калиди маҳсулотро нависед:\n\n" + text,
        parse_mode="HTML"
    )

    await state.set_state(AdminStates.waiting_delete_product)


@dp.message(AdminStates.waiting_delete_product, F.text)
async def delete_product_finish(message: Message, state: FSMContext):
    key = (message.text or "").strip()
    deleted = await db_delete_product(key)
    await state.clear()

    if deleted:
        await message.answer(
            "✅ Маҳсулот бо муваффақият нест карда шуд.",
            reply_markup=admin_kb
        )
    else:
        await message.answer(
            "❌ Ин маҳсулот ёфт нашуд.",
            reply_markup=admin_kb
        )

# ===================== CART =====================

@dp.message(F.text == "🛒 Сабад")
async def show_cart(message: Message) -> None:
    user = message.from_user
    if user is None:
        return
    items = await db_cart_get(user.id)
    if not items:
        empty_text = await db_get_text("cart_empty") or "🛒 Сабад холӣ аст."
        await message.answer(empty_text)
        return
    body, total = _cart_text(items)
    text = (
        "🛒 <b>Сабади шумо:</b>\n\n"
        + body
        + f"\n\n💵 <b>Ҷамъ: {total:.2f} сомонӣ</b>"
    )
    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Фармоиш додан", callback_data="send_order"),
                InlineKeyboardButton(text="🗑 Тоза кардан",   callback_data="clear_cart"),
            ]
        ]
    )
    await message.answer(text, parse_mode="HTML", reply_markup=kb)


@dp.message(Command("clearcart"))
async def clearcart_command(message: Message) -> None:
    user = message.from_user
    if user is None:
        return
    await db_cart_clear(user.id)
    await message.answer("🛒 Сабади шумо тоза шуд.")


@dp.message(F.text == "📞 Алоқа")
async def contact(message: Message) -> None:
    await message.answer(await db_get_text("contact"))


@dp.message(F.text == "🎁 Тахфиф")
async def discount_handler(message: Message) -> None:
    await message.answer(await db_get_text("discount"))

# ===================== DEBT CHECK (user) =====================

@dp.message(Command("debt"))
async def check_debt(message: Message) -> None:
    user = message.from_user
    if user is None:
        return
    args = (message.text or "").split(maxsplit=1)
    if len(args) < 2:
        await message.answer("ℹ️ Истифода: /debt <рақами телефон>")
        return
    phone = args[1].strip()
    debt  = await db_get_debt(phone)
    if debt:
        await message.answer(
            f"💳 <b>Маълумоти қарз</b>\n\n"
            f"📱 {phone}\n"
            f"💰 Қарз: {float(debt['amount']):.2f} сомонӣ\n"  # type: ignore
            f"📅 Санаи гирифтани қарз: {debt['created_at']}",
            parse_mode="HTML",
        )
    else:
        await message.answer("✅ Барои ин рақам қарз вуҷуд надорад.")

# ===================== ADMIN BUTTONS SET =====================

# FIX #5: Added the missing "📨 Хабардиҳӣ" button to the set
ADMIN_BUTTONS = {
    "🗑 Нархҳоро нишон диҳед",
    "➕ Маҳсулоти нав",
    "📥 Аз файл илова кун",
    "💰 Иваз кардани нарх",
    "📝 Иваз кардани матн",
    "💼 Иваз кардани расм",
    "➕ Иловаи қарз",
    "❌ Кам кардани қарз",
    "📂 Идораи категорияхо",
    "📨 Хабардиҳӣ",
}

@dp.message(F.text.in_(ADMIN_BUTTONS))
async def cancel_fsm_on_button(message: Message, state: FSMContext) -> None:
    await state.clear()
    txt = message.text
    if txt == "🗑 Нархҳоро нишон диҳед":
        await show_prices(message)
    elif txt == "➕ Маҳсулоти нав":
        await new_product_start(message, state)
    elif txt == "📥 Аз файл илова кун":
        await import_start(message, state)
    elif txt == "💰 Иваз кардани нарх":
        await change_price_start(message, state)
    elif txt == "📝 Иваз кардани матн":
        await change_text_start(message, state)
    elif txt == "💼 Иваз кардани расм":
        await change_photo_start(message, state)
    elif txt == "➕ Иловаи қарз":
        await new_debt(message, state)
    elif txt == "❌ Кам кардани қарз":
        await reduce_debt_start(message, state)
    elif txt == "📂 Идораи категорияхо":
        await category_menu(message, state)
    # FIX #6: Route broadcast button to its handler
    elif txt == "📨 Хабардиҳӣ":
        await broadcast_start(message, state)

# ─── Helper ───────────────────────────────────────────────────────────────────

async def _send_chunked(message: Message, header: str, lines: list[str]) -> None:
    chunk: list[str] = []
    size = 0
    for line in lines:
        if size + len(line) + 1 > 3800:
            await message.answer(header + "\n\n" + "\n".join(chunk), parse_mode="HTML")
            chunk, size = [], 0
        chunk.append(line)
        size += len(line) + 1
    if chunk:
        await message.answer(header + "\n\n" + "\n".join(chunk), parse_mode="HTML")

# ─── Show prices ──────────────────────────────────────────────────────────────

@dp.message(F.text == "🗑 Нархҳоро нишон диҳед")
async def show_prices(message: Message) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    prods = await db_all_products()
    if not prods:
        await message.answer("📦 Маҳсулот вуҷуд надорад.")
        return
    lines = [
        f"• <code>{p['key']}</code> [{p.get('category')}] — {p['name']}: {p['price']} сомонӣ"
        + (" 🖼" if p.get("photo_id") else "")
        for p in prods
    ]
    await _send_chunked(message, "📋 <b>Рӯйхати нархҳо:</b>", lines)

# ===================== BROADCAST =====================

# FIX #6: Implemented the missing broadcast feature end-to-end
@dp.message(F.text == "📨 Хабардиҳӣ")
async def broadcast_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    await message.answer(
        "📨 Матни хабарро нависед.\n"
        "Ин паём ба ҳамаи корбарони сабтшуда фиристода мешавад."
    )
    await state.set_state(AdminStates.waiting_broadcast_text)


@dp.message(AdminStates.waiting_broadcast_text, F.text)
async def broadcast_send(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        await state.clear()
        return

    text = (message.text or "").strip()
    if not text:
        await message.answer("❌ Матн холӣ аст. Дубора нависед:")
        return

    await state.clear()
    processing = await message.answer("⏳ Хабар фиристода мешавад...")

    telegram_ids = await db_get_all_telegram_ids()
    sent = failed = 0
    for tid in telegram_ids:
        try:
            await bot.send_message(tid, text)
            sent += 1
        except Exception:
            failed += 1

    await processing.delete()
    await message.answer(
        f"✅ Хабардиҳӣ анҷом ёфт.\n\n"
        f"📤 Фиристода шуд: <b>{sent}</b>\n"
        f"❌ Хато: <b>{failed}</b>",
        parse_mode="HTML",
        reply_markup=admin_kb,
    )

# ===================== CATEGORY MANAGEMENT =====================

async def category_menu(message: Message, state: FSMContext) -> None:
    """Open the category management sub-menu."""
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    cats = await db_get_categories()
    cats_list = ", ".join(f"<code>{c}</code>" for c in cats)
    await message.answer(
        f"📂 <b>Идораи категорияҳо</b>\n\n"
        f"Категорияҳои мавҷуда: {cats_list}\n\n"
        "Амалро интихоб кунед:",
        parse_mode="HTML",
        reply_markup=category_manage_kb,
    )


@dp.message(F.text == "📋 Рӯйхати категорияҳо")
async def list_categories(message: Message) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    cats = await db_get_categories()
    if not cats:
        await message.answer("⚠️ Категория вуҷуд надорад.")
        return
    lines = [f"{i+1}. <code>{c}</code>" for i, c in enumerate(cats)]
    await message.answer(
        "📋 <b>Категорияҳо:</b>\n\n" + "\n".join(lines),
        parse_mode="HTML",
    )


@dp.message(F.text == "➕ Илова кардани категория")
async def add_category_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    await message.answer(
        "✏️ Номи категорияи навро нависед\n"
        "(танҳо ҳарфҳои лотинӣ ва <code>_</code>):",
        parse_mode="HTML",
    )
    await state.set_state(AdminStates.waiting_new_category_name)


@dp.message(AdminStates.waiting_new_category_name, F.text)
async def add_category_receive(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        await state.clear()
        return
    raw = (message.text or "").strip()
    if not raw.replace("_", "").isalnum() or not raw.isascii():
        await message.answer(
            "❌ Ном бояд танҳо аз ҳарфҳои лотинӣ ва <code>_</code> иборат бошад. Дубора нависед:",
            parse_mode="HTML",
        )
        return
    name     = raw.lower()
    inserted = add_category_sync(name)
    await state.clear()
    if inserted:
        await message.answer(
            f"✅ Категорияи <code>{name}</code> бо муваффақият илова шуд!",
            parse_mode="HTML",
            reply_markup=category_manage_kb,
        )
    else:
        await message.answer(
            f"⚠️ Категорияи <code>{name}</code> аллакай мавҷуд аст.",
            parse_mode="HTML",
            reply_markup=category_manage_kb,
        )


@dp.message(F.text == "✏️ Тағйири номи категория")
async def rename_category_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    cats = await db_get_categories()
    lines = "\n".join(f"• <code>{c}</code>" for c in cats)
    await message.answer(
        f"📋 <b>Категорияҳо:</b>\n\n{lines}\n\n"
        "✏️ Номи категорияеро, ки тағйир медиҳед, нависед:",
        parse_mode="HTML",
    )
    await state.set_state(AdminStates.waiting_category_rename_select)


@dp.message(AdminStates.waiting_category_rename_select, F.text)
async def rename_category_select(message: Message, state: FSMContext) -> None:
    old_name = (message.text or "").strip().lower()
    cats     = await db_get_categories()
    if old_name not in cats:
        await message.answer(
            f"❌ Категорияи <code>{old_name}</code> ёфт нашуд. Дубора нависед:",
            parse_mode="HTML",
        )
        return
    await state.update_data(rename_old=old_name)
    await message.answer(
        f"✏️ Номи навро барои <code>{old_name}</code> нависед\n"
        "(танҳо ҳарфҳои лотинӣ ва <code>_</code>):",
        parse_mode="HTML",
    )
    await state.set_state(AdminStates.waiting_category_rename_new)


@dp.message(AdminStates.waiting_category_rename_new, F.text)
async def rename_category_new(message: Message, state: FSMContext) -> None:
    raw = (message.text or "").strip()
    if not raw.replace("_", "").isalnum() or not raw.isascii():
        await message.answer(
            "❌ Ном бояд танҳо аз ҳарфҳои лотинӣ ва <code>_</code> иборат бошад. Дубора нависед:",
            parse_mode="HTML",
        )
        return
    new_name = raw.lower()
    data     = await state.get_data()
    old_name = str(data["rename_old"])
    success  = await db_rename_category(old_name, new_name)
    await state.clear()
    if success:
        await message.answer(
            f"✅ Категория аз <code>{old_name}</code> ба <code>{new_name}</code> тағйир ёфт.",
            parse_mode="HTML",
            reply_markup=category_manage_kb,
        )
    else:
        await message.answer(
            f"⚠️ Категорияи <code>{new_name}</code> аллакай мавҷуд аст.",
            parse_mode="HTML",
            reply_markup=category_manage_kb,
        )


@dp.message(F.text == "🗑 Нест кардани категория")
async def delete_category_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    cats = await db_get_categories()
    deletable = [c for c in cats if c != "other"]
    if not deletable:
        await message.answer("⚠️ Нест кардан мумкин нест: танҳо «other» мавҷуд аст.")
        return
    lines = "\n".join(f"• <code>{c}</code>" for c in deletable)
    await message.answer(
        f"📋 <b>Категорияҳои нест карданшаванда:</b>\n\n{lines}\n\n"
        "✏️ Номи категорияро нависед:",
        parse_mode="HTML",
    )
    await state.set_state(AdminStates.waiting_category_delete_select)


@dp.message(AdminStates.waiting_category_delete_select, F.text)
async def delete_category_select(message: Message, state: FSMContext) -> None:
    name = (message.text or "").strip().lower()
    if name == "other":
        await message.answer("❌ Категорияи «other»-ро нест кардан мумкин нест.")
        return
    cats = await db_get_categories()
    if name not in cats:
        await message.answer(
            f"❌ Категорияи <code>{name}</code> ёфт нашуд. Дубора нависед:",
            parse_mode="HTML",
        )
        return
    await state.update_data(delete_name=name)
    await message.answer(
        f"⚠️ Шумо мехоҳед категорияи <code>{name}</code>-ро нест кунед?\n"
        "Маҳсулотҳо ба «other» мегузаранд.\n\n"
        "Барои тасдиқ <b>ха</b> нависед:",
        parse_mode="HTML",
    )
    await state.set_state(AdminStates.waiting_category_delete_confirm)


@dp.message(AdminStates.waiting_category_delete_confirm, F.text)
async def delete_category_confirm(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip().lower()
    data = await state.get_data()
    name = str(data["delete_name"])
    if text != "ха":
        await state.clear()
        await message.answer("❌ Нест кардан бекор шуд.", reply_markup=category_manage_kb)
        return
    await db_delete_category(name)
    await state.clear()
    await message.answer(
        f"✅ Категорияи <code>{name}</code> нест карда шуд.",
        parse_mode="HTML",
        reply_markup=category_manage_kb,
    )

# ─── Add new product ──────────────────────────────────────────────────────────

@dp.message(F.text == "➕ Маҳсулоти нав")
async def new_product_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    cats      = get_categories()
    cat_lines = "\n".join(f"• <code>{c}</code>" for c in cats)
    await message.answer(
        f"📦 <b>Категорияро интихоб кунед ё номи навро нависед:</b>\n\n"
        f"{cat_lines}\n\n"
        "💡 Агар категория мавҷуд набошад — номи навро нависед, худкор илова мешавад.",
        parse_mode="HTML",
        reply_markup=build_category_kb(),
    )
    await state.set_state(AdminStates.waiting_new_product_category)


@dp.message(AdminStates.waiting_new_product_category, F.text)
async def new_product_category(message: Message, state: FSMContext) -> None:
    raw = (message.text or "").strip()
    cat = raw.lower()
    if not cat.replace("_", "").isalnum() or not cat.isascii():
        await message.answer(
            "❌ Ном бояд танҳо аз ҳарфҳои лотинӣ ва <code>_</code> иборат бошад. Дубора нависед:",
            parse_mode="HTML",
        )
        return
    valid = set(get_categories())
    if cat not in valid:
        add_category_sync(cat)
        await message.answer(
            f"✅ Категорияи нав <code>{cat}</code> илова шуд!",
            parse_mode="HTML",
        )
    await state.update_data(new_category=cat)
    await message.answer(
        "✏️ Калиди маҳсулотро нависед.\n"
        "Танҳо ҳарфҳои лотинӣ ва <code>_</code>.\n"
        "Намуна: <code>sprite</code>",
        parse_mode="HTML",
        reply_markup=admin_kb,
    )
    await state.set_state(AdminStates.waiting_new_product_key)


@dp.message(AdminStates.waiting_new_product_key, F.text)
async def new_product_key(message: Message, state: FSMContext) -> None:
    raw = (message.text or "").strip()
    if not raw.replace("_", "").isalnum() or not raw.isascii():
        await message.answer(
            "❌ Калид бояд танҳо аз ҳарфҳои лотинӣ ва <code>_</code> иборат бошад. Дубора нависед:",
            parse_mode="HTML",
        )
        return
    key = raw.lower().replace(" ", "_")
    if await db_get_product(key):
        await message.answer(
            f"❌ Калиди <code>{key}</code> аллакай мавҷуд аст. Дигар калид нависед:",
            parse_mode="HTML",
        )
        return
    await state.update_data(new_key=key)
    await message.answer("✏️ Номи намоишии маҳсулотро нависед:")
    await state.set_state(AdminStates.waiting_new_product_name)


@dp.message(AdminStates.waiting_new_product_name, F.text)
async def new_product_name(message: Message, state: FSMContext) -> None:
    await state.update_data(new_name=(message.text or "").strip())
    await message.answer("💰 Нархро нависед (сомонӣ):")
    await state.set_state(AdminStates.waiting_new_product_price)


@dp.message(AdminStates.waiting_new_product_price, F.text)
async def new_product_price(message: Message, state: FSMContext) -> None:
    try:
        price = float((message.text or "").strip())
        if price < 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Лутфан рақами мусбат нависед:")
        return
    await state.update_data(new_price=price)
    await message.answer(
        "📸 Акси маҳсулотро фиристед.\n"
        "Агар акс нахоҳед — «бе акс» нависед."
    )
    await state.set_state(AdminStates.waiting_new_product_photo)


@dp.message(AdminStates.waiting_new_product_photo, F.photo)
async def new_product_photo(message: Message, state: FSMContext) -> None:
    fsm      = await state.get_data()
    key      = str(fsm["new_key"])
    name     = str(fsm["new_name"])
    price    = float(fsm["new_price"])  # type: ignore[arg-type]
    category = str(fsm.get("new_category"))
    photo_id = message.photo[-1].file_id  # type: ignore[index]
    await db_upsert_product(key, name, price, photo_id, category)
    await state.clear()
    await message.answer_photo(
        photo=photo_id,
        caption=(
            f"✅ Маҳсулоти нав илова шуд:\n\n"
            f"🔑 Калид: <code>{key}</code>\n"
            f"📦 Категория: <b>{category}</b>\n"
            f"🛍 Ном: <b>{name}</b>\n"
            f"💰 Нарх: <b>{price} сомонӣ</b>"
        ),
        parse_mode="HTML",
    )
    await message.answer("👨‍💼 Панели админ:", reply_markup=admin_kb)


@dp.message(AdminStates.waiting_new_product_photo, F.text == "бе акс")
async def new_product_photo_skip(message: Message, state: FSMContext) -> None:
    fsm      = await state.get_data()
    key      = str(fsm["new_key"])
    name     = str(fsm["new_name"])
    price    = float(fsm["new_price"])  # type: ignore[arg-type]
    category = str(fsm.get("new_category"))
    await db_upsert_product(key, name, price, category=category)
    await state.clear()
    await message.answer(
        f"✅ Маҳсулоти нав (бе акс) илова шуд:\n\n"
        f"🔑 Калид: <code>{key}</code>\n"
        f"📦 Категория: <b>{category}</b>\n"
        f"🛍 Ном: <b>{name}</b>\n"
        f"💰 Нарх: <b>{price} сомонӣ</b>",
        parse_mode="HTML",
        reply_markup=admin_kb,
    )


@dp.message(AdminStates.waiting_new_product_photo)
async def new_product_photo_wrong(message: Message) -> None:
    await message.answer("📸 Лутфан акс фиристед ё «бе акс» нависед.")

# ─── Import from file ─────────────────────────────────────────────────────────

@dp.message(F.text == "📥 Аз файл илова кун")
async def import_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    cats_str = " | ".join(get_categories())
    await message.answer(
        "📥 Файли маҳсулотҳоро бифиристед.\n\n"
        "Форматҳои дастгирӣшаванда: <b>.csv</b> ё <b>.xlsx</b>\n\n"
        "Сутунҳои лозимӣ:\n"
        "• <code>key</code>      — калиди ягона (лотинӣ)\n"
        "• <code>name</code>     — номи маҳсулот\n"
        "• <code>price</code>    — нарх (рақам)\n\n"
        "Сутунҳои ихтиёрӣ:\n"
        f"• <code>category</code> — {cats_str}\n"
        "• <code>photo_id</code> — Telegram file_id-и акс",
        parse_mode="HTML",
    )
    await state.set_state(AdminStates.waiting_file_import)


@dp.message(AdminStates.waiting_file_import, F.document)
async def import_file_receive(message: Message, state: FSMContext) -> None:
    doc = message.document
    if doc is None:
        await message.answer("📎 Лутфан файли .csv ё .xlsx фиристед.")
        return
    filename = doc.file_name or ""
    if not filename.endswith((".csv", ".xlsx", ".xls")):
        await message.answer("❌ Формати файл дуруст нест. Танҳо .csv ё .xlsx фиристед.")
        return
    processing = await message.answer("⏳ Файл коркард мешавад...")
    tg_file    = await bot.get_file(doc.file_id)
    file_path  = tg_file.file_path
    if file_path is None:
        await processing.delete()
        await message.answer("❌ Файлро боргирӣ кардан имконнопазир аст.")
        await state.clear()
        return
    dl = await bot.download_file(file_path)
    if dl is None:
        await processing.delete()
        await message.answer("❌ Боргирии файл нокомёб шуд.")
        await state.clear()
        return
    raw_bytes = dl.read()
    result    = parse_import_file(raw_bytes, filename)
    if isinstance(result, str):
        await processing.delete()
        await message.answer(result, reply_markup=admin_kb)
        await state.clear()
        return
    added = updated = 0
    for p in result:
        existing = await db_get_product(str(p["key"]))
        await db_upsert_product(
            str(p["key"]),
            str(p["name"]),
            float(p["price"]),  # type: ignore[arg-type]
            str(p["photo_id"]) if p.get("photo_id") else None,
            str(p.get("category") or "other"),
        )
        if existing:
            updated += 1
        else:
            added += 1
    await state.clear()
    await processing.delete()
    await message.answer(
        f"✅ Файл бо муваффақият коркард шуд!\n\n"
        f"➕ Нав илова шуд: <b>{added}</b>\n"
        f"🔄 Навсозӣ шуд: <b>{updated}</b>\n"
        f"📦 Ҷамъ: <b>{added + updated}</b> маҳсулот",
        parse_mode="HTML",
        reply_markup=admin_kb,
    )


@dp.message(AdminStates.waiting_file_import)
async def import_wrong_type(message: Message) -> None:
    await message.answer("📎 Лутфан файли .csv ё .xlsx фиристед.")

# ─── Change price ─────────────────────────────────────────────────────────────

@dp.message(F.text == "💰 Иваз кардани нарх")
async def change_price_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    prods = await db_all_products()
    if not prods:
        await message.answer("❌ Дар база ягон маҳсулот вуҷуд надорад.")
        return
    lines = [
        f"• <code>{p['key']}</code> — {p['name']}: {p['price']} сомонӣ"
        for p in prods
    ]
    await _send_chunked(message, "📋 <b>Маҳсулотҳо:</b>", lines)
    await message.answer("✏️ Калиди маҳсулотро нависед:")
    await state.set_state(AdminStates.waiting_price_key)


@dp.message(AdminStates.waiting_price_key, F.text)
async def change_price_key(message: Message, state: FSMContext) -> None:
    key     = (message.text or "").strip()
    product = await db_get_product(key)
    if not product:
        await message.answer(
            f"❌ Калид <code>{key}</code> ёфт нашуд. Дубора нависед:", parse_mode="HTML"
        )
        return
    await state.update_data(price_key=key)
    await message.answer(
        f"💰 Нархи нави <b>{product['name']}</b>-ро нависед:", parse_mode="HTML"
    )
    await state.set_state(AdminStates.waiting_price_value)


@dp.message(AdminStates.waiting_price_value, F.text)
async def change_price_value(message: Message, state: FSMContext) -> None:
    try:
        new_price = float((message.text or "").strip())
        if new_price < 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Лутфан рақами мусбат нависед:")
        return
    fsm_data = await state.get_data()
    key      = str(fsm_data["price_key"])
    product  = await db_get_product(key)
    if not product:
        await message.answer("❌ Маҳсулот ёфт нашуд.")
        await state.clear()
        return
    existing_photo    = str(product["photo_id"]) if product.get("photo_id") else None
    existing_category = str(product.get("category"))
    await db_upsert_product(key, str(product["name"]), new_price, existing_photo, existing_category)
    await state.clear()
    await message.answer(
        f"✅ Нархи <b>{product['name']}</b> ба <b>{new_price} сомонӣ</b> иваз шуд.",
        parse_mode="HTML",
        reply_markup=admin_kb,
    )

# ─── Change text ──────────────────────────────────────────────────────────────

@dp.message(F.text == "📝 Иваз кардани матн")
async def change_text_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    texts = await db_all_texts()
    lines = [f"• <code>{t['key']}</code>" for t in texts]
    await _send_chunked(message, "📋 <b>Матнҳо:</b>", lines)
    await message.answer("✏️ Калидро нависед:")
    await state.set_state(AdminStates.waiting_text_key)


@dp.message(AdminStates.waiting_text_key, F.text)
async def change_text_key(message: Message, state: FSMContext) -> None:
    key   = (message.text or "").strip()
    texts = await db_all_texts()
    if key not in {str(t["key"]) for t in texts}:
        await message.answer(
            f"❌ Калид <code>{key}</code> ёфт нашуд. Дубора нависед:", parse_mode="HTML"
        )
        return
    value = await db_get_text(key)
    await state.update_data(text_key=key)
    await message.answer(f"📝 Матни кунунӣ:\n\n{value}\n\n✏️ Матни навро нависед:")
    await state.set_state(AdminStates.waiting_text_value)


@dp.message(AdminStates.waiting_text_value, F.text)
async def change_text_value(message: Message, state: FSMContext) -> None:
    fsm_data = await state.get_data()
    key      = str(fsm_data["text_key"])
    await db_set_text(key, (message.text or "").strip())
    await state.clear()
    await message.answer(
        f"✅ Матни <code>{key}</code> бо муваффақият навсозӣ шуд.",
        parse_mode="HTML",
        reply_markup=admin_kb,
    )

# ─── Change photo ─────────────────────────────────────────────────────────────

@dp.message(F.text == "💼 Иваз кардани расм")
async def change_photo_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    prods = await db_all_products()
    if not prods:
        await message.answer("📦 Ҳоло ягон маҳсулот вуҷуд надорад.")
        return
    lines = [
        f"• <code>{p['key']}</code> — {p['name']}"
        for p in prods
    ]
    await _send_chunked(message, "📋 <b>Маҳсулотҳо (KEY):</b>", lines)
    await message.answer("✏️ Калиди маҳсулотро нависед:")
    await state.set_state(AdminStates.waiting_photo_key)


@dp.message(AdminStates.waiting_photo_key, F.text)
async def change_photo_get_key(message: Message, state: FSMContext) -> None:
    key     = (message.text or "").strip()
    product = await db_get_product(key)
    if not product:
        await message.answer("❌ Ин калид ёфт нашуд, дубора нависед:")
        return
    await state.update_data(photo_key=key)
    current_photo = product.get("photo_id")
    if current_photo:
        await message.answer_photo(
            photo=str(current_photo),
            caption=f"🖼 Расми кунунии <b>{product['name']}</b>\n\n📸 Расми навро фиристед:",
            parse_mode="HTML",
        )
    else:
        await message.answer(
            f"📦 <b>{product['name']}</b> ҳоло расм надорад.\n\n📸 Расми навро фиристед:",
            parse_mode="HTML",
        )
    await state.set_state(AdminStates.waiting_photo_new)


@dp.message(AdminStates.waiting_photo_new, F.photo)
async def change_photo_get_new(message: Message, state: FSMContext) -> None:
    fsm_data = await state.get_data()
    key      = str(fsm_data["photo_key"])
    photo_id = message.photo[-1].file_id  # type: ignore[index]
    await db_set_product_photo(key, photo_id)
    await state.clear()
    await message.answer_photo(
        photo=photo_id,
        caption="✅ Расми маҳсулот бо муваффақият иваз шуд!\n\nАкнун корбарон расми навро мебинанд.",
    )
    await message.answer("👨‍💼 Панели админ:", reply_markup=admin_kb)


@dp.message(AdminStates.waiting_photo_new)
async def change_photo_wrong(message: Message) -> None:
    await message.answer("📸 Лутфан танҳо акс фиристед.")

# ─── Add debt ─────────────────────────────────────────────────────────────────

@dp.message(F.text == "➕ Иловаи қарз")
async def new_debt(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    await message.answer("📱 Рақами телефони харидорро нависед:")
    await state.set_state(DebtState.phone)


@dp.message(DebtState.phone, F.text)
async def debt_got_phone(message: Message, state: FSMContext) -> None:
    phone      = (message.text or "").strip()
    await state.update_data(phone=phone)
    cart_items = await db_cart_get_by_phone(phone)
    if cart_items:
        body, total = _cart_text(cart_items)
        await state.update_data(cart_total=total)
        existing = await db_get_debt(phone)
        prev = (
            f"\n⚠️ Қарзи қаблӣ: <b>{float(existing['amount']):.2f} сомонӣ</b>"  # type: ignore[index]
            if existing else ""
        )
        await message.answer(
            f"🛒 <b>Сабади {phone}:</b>\n\n"
            + body
            + f"\n\n💵 <b>Ҷамъ: {total:.2f} сомонӣ</b>"
            + prev
            + "\n\nБарои қарз кардани ин маблағ <b>ха</b> нависед ё маблағи дигарро нависед:",
            parse_mode="HTML",
        )
        await state.set_state(DebtState.confirm)
    else:
        existing = await db_get_debt(phone)
        if existing:
            await message.answer(
                f"ℹ️ Қарзи қаблӣ барои <code>{phone}</code>: "
                f"<b>{float(existing['amount']):.2f} сомонӣ</b>\n\n"  # type: ignore[index]
                "💰 Маблағи қарзи навро нависед (ба қаблӣ илова мешавад):",
                parse_mode="HTML",
            )
        else:
            await message.answer(
                f"ℹ️ Сабади {phone} холӣ аст ё регистратсия нашудааст.\n\n"
                "💰 Маблағи қарзро дастӣ нависед:"
            )
        await state.set_state(DebtState.amount)


@dp.message(DebtState.confirm, F.text)
async def debt_confirm(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip().lower()
    data = await state.get_data()
    if text == "ха":
        amount = float(data.get("cart_total", 0.0))
    else:
        try:
            amount = float(text)
            if amount <= 0:
                raise ValueError
        except ValueError:
            await message.answer("❌ Лутфан рақами мусбат ё «ха» нависед:")
            return
    await _record_debt(message, state, str(data["phone"]), amount)


@dp.message(DebtState.amount, F.text)
async def debt_manual_amount(message: Message, state: FSMContext) -> None:
    text = (message.text or "").strip()
    try:
        amount = float(text)
        if amount <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Лутфан рақами мусбат нависед:")
        return
    data = await state.get_data()
    await _record_debt(message, state, str(data["phone"]), amount)


async def _record_debt(message: Message, state: FSMContext, phone: str, amount: float) -> None:
    existing = await db_get_debt(phone)
    if existing:
        new_amount = float(existing["amount"]) + amount  # type: ignore[index]
        await db_update_debt(phone, new_amount)
    else:
        new_amount = amount
        await db_create_debt(phone, amount)
    telegram_id = await db_get_user_telegram_id(phone)
    if telegram_id:
        await db_cart_clear(telegram_id)
    await state.clear()
    await message.answer(
        f"✅ Қарз сабт шуд.\n\n"
        f"📱 {phone}\n"
        f"💸 Илова шуд: <b>{amount:.2f} сомонӣ</b>\n"
        f"💰 Ҷамъи қарз: <b>{new_amount:.2f} сомонӣ</b>",
        parse_mode="HTML",
        reply_markup=admin_kb,
    )
    if telegram_id:
        await _notify_user(
            telegram_id,
            f"💳 Ба ҳисоби шумо қарз илова шуд.\n\n"
            f"💸 Қарзи нав: {amount:.2f} сомонӣ\n"
            f"💰 Ҷамъ: {new_amount:.2f} сомонӣ",
        )

# ─── Reduce / clear debt ──────────────────────────────────────────────────────

@dp.message(F.text == "❌ Кам кардани қарз")
async def reduce_debt_start(message: Message, state: FSMContext) -> None:
    user = message.from_user
    if user is None or not is_admin(user.id):
        return
    await message.answer("📱 Рақами телефонро нависед:")
    await state.set_state(DeleteDebt.phone)


@dp.message(DeleteDebt.phone, F.text)
async def reduce_debt_phone(message: Message, state: FSMContext) -> None:
    phone = (message.text or "").strip()
    debt  = await db_get_debt(phone)
    if not debt:
        await message.answer("✅ Барои ин рақам қарз вуҷуд надорад.", reply_markup=admin_kb)
        await state.clear()
        return
    await state.update_data(phone=phone)
    await message.answer(
        f"💳 Қарзи ҷорӣ барои <code>{phone}</code>: "
        f"<b>{float(debt['amount']):.2f} сомонӣ</b>\n\n"  # type: ignore[index]
        "Чанд сомонӣ пардохт кард?\n"
        "(Барои нест кардани пурра <b>ҳама</b> нависед)",
        parse_mode="HTML",
    )
    await state.set_state(DeleteDebt.amount)


@dp.message(DeleteDebt.amount, F.text)
async def reduce_debt_amount(message: Message, state: FSMContext) -> None:
    text  = (message.text or "").strip().lower()
    data  = await state.get_data()
    phone = str(data["phone"])
    debt  = await db_get_debt(phone)
    if not debt:
        await message.answer("❌ Қарз ёфт нашуд.")
        await state.clear()
        return
    if text == "ҳама":
        paid      = float(debt["amount"])  # type: ignore[index]
        remaining = 0.0
    else:
        try:
            paid = float(text)
            if paid <= 0:
                raise ValueError
        except ValueError:
            await message.answer("❌ Лутфан рақами мусбат ё «ҳама» нависед:")
            return
        remaining = max(0.0, float(debt["amount"]) - paid)  # type: ignore[index]
    if remaining <= 0:
        await db_delete_debt(phone)
        reply = f"✅ Қарзи <code>{phone}</code> пурра пардохт шуд ва нест карда шуд."
    else:
        await db_update_debt(phone, remaining)
        reply = (
            f"✅ Пардохт қабул шуд.\n\n"
            f"📱 {phone}\n"
            f"💸 Пардохт шуд: <b>{paid:.2f} сомонӣ</b>\n"
            f"💰 Боқимонда: <b>{remaining:.2f} сомонӣ</b>"
        )
    await state.clear()
    await message.answer(reply, parse_mode="HTML", reply_markup=admin_kb)
    telegram_id = await db_get_user_telegram_id(phone)
    if telegram_id:
        notif = (
            f"✅ Пардохти қарзи шумо қабул шуд.\n\n"
            f"💸 Пардохт шуд: {paid:.2f} сомонӣ\n"
            + (f"💰 Боқимонда: {remaining:.2f} сомонӣ"
               if remaining > 0 else "🎉 Қарзи шумо пурра адо шуд!")
        )
        await _notify_user(telegram_id, notif)

# ─── Catch-all for dynamic categories (must be last) ─────────────────────────

@dp.message(F.text)
async def dynamic_category_handler(message: Message, state: FSMContext) -> None:
    """Handle button presses for any category added by the admin at runtime."""
    text = (message.text or "").strip()
    current_state = await state.get_state()
    if current_state is not None:
        return
    cat_key = _cat_key_from_label(text)
    if cat_key is None:
        return
    await _send_category_menu(message, cat_key, f"📦 {text}")

# ===================== MAIN =====================

if __name__ == "__main__":
    async def main() -> None:
        await init_db()
        print("Bot started")
        await dp.start_polling(bot)

    asyncio.run(main())
    